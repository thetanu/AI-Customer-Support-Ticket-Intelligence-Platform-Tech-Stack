import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference

# Define file paths
csv_path = r"c:\Users\ntanu\OneDrive\Desktop\AI-Customer-Support-Ticket-Intelligence-Platform\dataset\cleaned_customer_support_tickets.csv"
excel_path = r"c:\Users\ntanu\OneDrive\Desktop\AI-Customer-Support-Ticket-Intelligence-Platform\excel\customer_support_tickets_analytics.xlsx"

# 1. Load data
print("Loading and cleaning data...")
df = pd.read_csv(csv_path)

# Deduplicate by Ticket ID if any duplicates exist
df = df.drop_duplicates(subset=["Ticket_ID"])

# Handle missing values
df["Resolution"] = df["Resolution"].fillna("Under Investigation")
df["Customer_Satisfaction_Rating"] = df["Customer_Satisfaction_Rating"].fillna(-1)
df["Time_to_Resolution"] = df["Time_to_Resolution"].fillna("")

# Extract Year-Month for trend analysis
df["Created_Date"] = pd.to_datetime(df["Created_Date"])
df["Year_Month"] = df["Created_Date"].dt.strftime("%Y-%m")

# Create workbook
wb = openpyxl.Workbook()

# Sheet 1: Executive Dashboard
ws_dash = wb.active
ws_dash.title = "Executive Dashboard"
ws_dash.views.sheetView[0].showGridLines = True

# Sheet 2: Cleaned Data
ws_data = wb.create_sheet(title="Cleaned Data")
ws_data.views.sheetView[0].showGridLines = True

# 2. Write Cleaned Data
print("Writing cleaned data to Excel...")
# Headers
headers = list(df.columns)
ws_data.append(headers)

# Rows
for r in df.values.tolist():
    ws_data.append(r)

# Auto-fit columns for Cleaned Data
for col in ws_data.columns:
    max_len = max(len(str(cell.value or '')) for cell in col)
    col_letter = get_column_letter(col[0].column)
    ws_data.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 30)

# 3. Design Executive Dashboard
print("Building Executive Dashboard...")
# Theme colors
NAVY_HEADER = "0F172A"
KPI_BG = "F1F5F9"
BORDER_COLOR = "CBD5E1"
ACCENT_BLUE = "2563EB"

font_title = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
font_section = Font(name="Segoe UI", size=13, bold=True, color="0F172A")
font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
font_bold = Font(name="Segoe UI", size=11, bold=True, color="0F172A")
font_normal = Font(name="Segoe UI", size=11, color="334155")
font_kpi_num = Font(name="Segoe UI", size=18, bold=True, color="1E3A8A")
font_kpi_label = Font(name="Segoe UI", size=9, bold=True, color="475569")

fill_title = PatternFill(fill_type="solid", start_color=NAVY_HEADER, end_color=NAVY_HEADER)
fill_kpi = PatternFill(fill_type="solid", start_color=KPI_BG, end_color=KPI_BG)
fill_header = PatternFill(fill_type="solid", start_color="1E40AF", end_color="1E40AF")
fill_total = PatternFill(fill_type="solid", start_color="E2E8F0", end_color="E2E8F0")

align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="center")
align_right = Alignment(horizontal="right", vertical="center")

thin_border_side = Side(border_style="thin", color=BORDER_COLOR)
thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
thick_bottom = Border(bottom=Side(border_style="medium", color="0F172A"))

# Title Banner
ws_dash.merge_cells("B2:N3")
title_cell = ws_dash["B2"]
title_cell.value = "AI Customer Support Ticket Intelligence Dashboard"
title_cell.font = font_title
title_cell.fill = fill_title
title_cell.alignment = align_center

# Apply styling to all cells in the title merge range to avoid display glitches
for r in range(2, 4):
    for c in range(2, 15):
        ws_dash.cell(row=r, column=c).fill = fill_title

# 4. KPI Cards Layout
kpis = [
    {"label": "TOTAL TICKETS", "formula": "=COUNTA('Cleaned Data'!A2:A8470)", "col_start": 2, "col_end": 3},
    {"label": "AVG RESOLUTION TIME (HRS)", "formula": "=ROUND(AVERAGE('Cleaned Data'!X2:X8470), 2)", "col_start": 4, "col_end": 5},
    {"label": "HIGH PRIORITY TICKETS", "formula": "=COUNTIF('Cleaned Data'!M2:M8470, \"High\")", "col_start": 6, "col_end": 7},
    {"label": "OPEN TICKETS", "formula": "=COUNTIF('Cleaned Data'!K2:K8470, \"Open\")", "col_start": 8, "col_end": 9},
    {"label": "CLOSED TICKETS", "formula": "=COUNTIF('Cleaned Data'!K2:K8470, \"Closed\")", "col_start": 10, "col_end": 11},
    {"label": "AVERAGE CSAT RATING", "formula": "=ROUND(AVERAGEIFS('Cleaned Data'!Q2:Q8470, 'Cleaned Data'!Q2:Q8470, \">=1\"), 2)", "col_start": 12, "col_end": 13}
]

for kpi in kpis:
    # Merge label cells
    ws_dash.merge_cells(start_row=5, start_column=kpi["col_start"], end_row=5, end_column=kpi["col_end"])
    lbl_cell = ws_dash.cell(row=5, column=kpi["col_start"])
    lbl_cell.value = kpi["label"]
    lbl_cell.font = font_kpi_label
    lbl_cell.fill = fill_kpi
    lbl_cell.alignment = align_center
    
    # Merge value cells
    ws_dash.merge_cells(start_row=6, start_column=kpi["col_start"], end_row=7, end_column=kpi["col_end"])
    val_cell = ws_dash.cell(row=6, column=kpi["col_start"])
    val_cell.value = kpi["formula"]
    val_cell.font = font_kpi_num
    val_cell.fill = fill_kpi
    val_cell.alignment = align_center

    # Apply borders and backgrounds to all merged cells
    for r in range(5, 8):
        for c in range(kpi["col_start"], kpi["col_end"] + 1):
            cell = ws_dash.cell(row=r, column=c)
            cell.fill = fill_kpi
            cell.border = thin_border

# 5. Table 1: Ticket Category Distribution by Priority
ws_dash.cell(row=9, column=2, value="Ticket Distribution by Category & Priority").font = font_section
ws_dash.merge_cells("B9:F9")

headers_cat = ["Ticket Category", "High", "Medium", "Low", "Total"]
for c_idx, h in enumerate(headers_cat, start=2):
    cell = ws_dash.cell(row=10, column=c_idx, value=h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center
    cell.border = thin_border

categories = ["Account", "Billing", "Login", "Refund", "Shipping", "Technical"]
for r_idx, cat in enumerate(categories, start=11):
    ws_dash.cell(row=r_idx, column=2, value=cat).font = font_bold
    ws_dash.cell(row=r_idx, column=2).border = thin_border
    
    # Formulas for High, Medium, Low
    for c_idx, priority in enumerate(["High", "Medium", "Low"], start=3):
        cell = ws_dash.cell(row=r_idx, column=c_idx)
        cell.value = f'=COUNTIFS(\'Cleaned Data\'!$T$2:$T$8470, B{r_idx}, \'Cleaned Data\'!$M$2:$M$8470, "{priority}")'
        cell.font = font_normal
        cell.alignment = align_center
        cell.border = thin_border
        
    # Total Formula
    tot_cell = ws_dash.cell(row=r_idx, column=6)
    tot_cell.value = f"=SUM(C{r_idx}:E{r_idx})"
    tot_cell.font = font_bold
    tot_cell.alignment = align_center
    tot_cell.border = thin_border

# Total Row
tot_row_idx = 17
ws_dash.cell(row=tot_row_idx, column=2, value="Total").font = font_bold
ws_dash.cell(row=tot_row_idx, column=2).fill = fill_total
ws_dash.cell(row=tot_row_idx, column=2).border = thin_border

for c_idx in range(3, 7):
    col_letter = get_column_letter(c_idx)
    cell = ws_dash.cell(row=tot_row_idx, column=c_idx)
    cell.value = f"=SUM({col_letter}11:{col_letter}16)"
    cell.font = font_bold
    cell.fill = fill_total
    cell.alignment = align_center
    cell.border = thin_border

# 6. Table 2: Top 10 Support Agents Performance
ws_dash.cell(row=19, column=2, value="Top 10 Support Agents Performance").font = font_section
ws_dash.merge_cells("B19:E19")

headers_agent = ["Support Agent", "Tickets Assigned", "Average CSAT", "Avg Resolution Time (hrs)"]
for c_idx, h in enumerate(headers_agent, start=2):
    cell = ws_dash.cell(row=20, column=c_idx, value=h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center
    cell.border = thin_border

# Get top 10 agents by ticket counts in Python first to populate their names
top_agents = df["Support_Agent"].value_counts().head(10).index.tolist()

for r_offset, agent in enumerate(top_agents):
    r_idx = 21 + r_offset
    ws_dash.cell(row=r_idx, column=2, value=agent).font = font_bold
    ws_dash.cell(row=r_idx, column=2).border = thin_border
    
    # Tickets Assigned
    t_cell = ws_dash.cell(row=r_idx, column=3)
    t_cell.value = f"=COUNTIF('Cleaned Data'!$U$2:$U$8470, B{r_idx})"
    t_cell.font = font_normal
    t_cell.alignment = align_center
    t_cell.border = thin_border
    
    # Average CSAT (ignoring -1)
    csat_cell = ws_dash.cell(row=r_idx, column=4)
    csat_cell.value = f'=ROUND(AVERAGEIFS(\'Cleaned Data\'!$Q$2:$Q$8470, \'Cleaned Data\'!$U$2:$U$8470, B{r_idx}, \'Cleaned Data\'!$Q$2:$Q$8470, ">=1"), 2)'
    csat_cell.font = font_normal
    csat_cell.alignment = align_center
    csat_cell.border = thin_border
    
    # Avg Resolution Time
    res_cell = ws_dash.cell(row=r_idx, column=5)
    res_cell.value = f"=ROUND(AVERAGEIF('Cleaned Data'!$U$2:$U$8470, B{r_idx}, 'Cleaned Data'!$X$2:$X$8470), 2)"
    res_cell.font = font_normal
    res_cell.alignment = align_center
    res_cell.border = thin_border

# 7. Table 3: Monthly Ticket Volume Trends
ws_dash.cell(row=33, column=2, value="Monthly Ticket Volume Trends").font = font_section
ws_dash.merge_cells("B33:C33")

headers_month = ["Month", "Tickets Logged"]
for c_idx, h in enumerate(headers_month, start=2):
    cell = ws_dash.cell(row=34, column=c_idx, value=h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center
    cell.border = thin_border

# Get sorted list of months in Python
months = sorted(df["Year_Month"].unique().tolist())

for r_offset, month in enumerate(months):
    r_idx = 35 + r_offset
    ws_dash.cell(row=r_idx, column=2, value=month).font = font_bold
    ws_dash.cell(row=r_idx, column=2).border = thin_border
    
    # COUNTIF formula based on Year_Month column (Column Y in Cleaned Data)
    m_cell = ws_dash.cell(row=r_idx, column=3)
    m_cell.value = f"=COUNTIF('Cleaned Data'!$Y$2:$Y$8470, B{r_idx})"
    m_cell.font = font_normal
    m_cell.alignment = align_center
    m_cell.border = thin_border

# 8. Add Charts
print("Adding Excel Charts...")
# Bar Chart for Category Distribution
bar_chart = BarChart()
bar_chart.type = "col"
bar_chart.style = 10
bar_chart.title = "Ticket Category Distribution by Priority"
bar_chart.y_axis.title = "Number of Tickets"
bar_chart.x_axis.title = "Category"
bar_chart.width = 16
bar_chart.height = 10

data_ref = Reference(ws_dash, min_col=3, min_row=10, max_col=5, max_row=16)
cats_ref = Reference(ws_dash, min_col=2, min_row=11, max_row=16)
bar_chart.add_data(data_ref, titles_from_data=True)
bar_chart.set_categories(cats_ref)
ws_dash.add_chart(bar_chart, "H9")

# Line Chart for Monthly Trends
line_chart = LineChart()
line_chart.title = "Monthly Ticket Volume Trends"
line_chart.style = 13
line_chart.y_axis.title = "Tickets"
line_chart.x_axis.title = "Month"
line_chart.width = 16
line_chart.height = 10

data_ref2 = Reference(ws_dash, min_col=3, min_row=34, max_row=34 + len(months))
cats_ref2 = Reference(ws_dash, min_col=2, min_row=35, max_row=34 + len(months))
line_chart.add_data(data_ref2, titles_from_data=True)
line_chart.set_categories(cats_ref2)
ws_dash.add_chart(line_chart, "H23")

# Adjust column widths on Dashboard
ws_dash.column_dimensions["A"].width = 3
ws_dash.column_dimensions["B"].width = 25
ws_dash.column_dimensions["C"].width = 15
ws_dash.column_dimensions["D"].width = 15
ws_dash.column_dimensions["E"].width = 15
ws_dash.column_dimensions["F"].width = 15
ws_dash.column_dimensions["G"].width = 3

# Save workbook
print(f"Saving workbook to {excel_path}...")
os.makedirs(os.path.dirname(excel_path), exist_ok=True)
wb.save(excel_path)
print("Workbook created successfully!")
wb.close()
